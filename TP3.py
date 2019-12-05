"KOUATELAY Anne Keren TP3 B1 "


# 2 Parseur HTML:BeautifulSoup

import sys
from bs4 import BeautifulSoup

with open ("mon_cv.html","r") as f:
    html_doc= f.read()

soup=BeautifulSoup (html_doc, "html.parser")

    
    
def find_name():
    print(soup.find(id="nom"))
    
def find_address():
    print (soup.find(id="adresse"))
    
    
def find_phone():
    print (soup.find(id="telephone"))
    
def find_mail():
    link=soup.find(id="email")
    print(link.get('href'))
    

    

find_name()
find_address()
find_phone()
find_mail()


#3 Manipulation d Excel en Python

import openpyxl
# Load in the workbook 
wb = openpyxl.load_workbook('cartes.xlsx')
# Get sheet names 
sheets = wb.sheetnames 
ws = wb[sheets[0]] 
for row in ws.iter_rows(min_row=3,values_only=True): 
    print(row[1],row[2])
    
#4 Importation des données à partir d un site web
    
import urllib.request
carte="Pikachu 58/102" 
carteprim= carte.replace(" ","%20")
url="https://www.ebay.fr/sch/i.html?_from=R40&_sacat=0&_nkw="+carteprim+"&LH_Complete=1&LH_Sold=1&rt=ncRemarquer"
f=urllib.request.urlopen(url)
html_doc=f.read()
soup=BeautifulSoup(html_doc,"html.parser")
for k in soup.find_all("li",_sp="p2045573.m1686.l0",attrs={"class":"sresult"}):
    print (k)
    print ("--------------------------------------------------------------------")


from openpyxl import Workbook


workbook=Workbook()
sheet= workbook.active
carte="Pikachu 58/102"
carteprim= carte.replace(" ","%20")
url="https://www.ebay.fr/sch/i.html?_from=R40&_sacat=0&_nkw="+carteprim+"&LH_Complete=1&LH_Sold=1&rt=ncRemarquer"
f=urllib.request.urlopen(url)
html_doc=f.read()
soup=BeautifulSoup(html_doc,"html.parser")
for k in soup.find_all("li",_sp="p2045573.m1686.l0",attrs={"class":"sresult"}):
    p=k.select("h3.lvtitle")[0]
    titre=p.select_one("a").get_text().strip()
    p=k.select("li.lvprice")[0]
    prix=p.select_one("span").get_text().strip()[:-4]
    p=k.select("ul.lvdetails")[0]
    lis=p.select("li")
    date=lis[0].get_text().strip()
    if len(lis)>1:
        result= lis[1].get_text().find('Provenance')
        if result>0:
            pays=(lis[1].get_text().strip())[13:]
        else:
            pays="France"
    print (titre,prix,date,pays)
    sheet.append([carte,titre,prix,date,pays])
    print("---------------------------------------------------------------------")
    
    workbook.save(filename="ventes.xlsx")
    
    

#Le dernier programme ne marche pas au niveau de l'affectation de "pays" 


















